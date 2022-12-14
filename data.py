import datetime
from time import strftime

from openpyxl import load_workbook
# lookFor = '990110162483870'
# def data_by_trek(lookFor):
#     for i in range(1, sheet.max_row+1):
#         value = sheet.cell(row=i, column=4).value
#         if value == lookFor:
#             availble = sheet[i][10].value
#             price = sheet[i][11].value
#             trek = sheet[i][3].value
#             position = sheet[i][5].value
#             return [trek, availble, position, price]

# print(data_by_trek(lookFor=lookFor))

# name2 = '294'
# name3 = 'GX-0294'


async def data_by_id(id1, id2, id3, id4):
    wb_search = load_workbook('data.xlsx')
    sheet = wb_search.active
    data = []
    for i in range(1, sheet.max_row + 1):
        data_dict = {}
        value = str(sheet.cell(row=i, column=3).value)
        if id1 == value or id2 == value or id3 == value or id4 == value:
            data_dict['no'] = id1
            data_dict['no_auto'] = sheet[i][0].value
            try:
                data_dict['date'] = sheet[i][10].value.strftime('%d.%m.%Y')
            except AttributeError:
                data_dict['date'] = None

            data_dict['trek'] = sheet[i][3].value

            data_dict['weight'] = sheet[i][7].value
            data_dict['price'] = sheet[i][8].value
            data_dict['in_china'] = sheet[i][6].value
            if data_dict.get('in_china') is None:
                data_dict['in_china'] = '0'
            if data_dict.get('weight') is None:
                data_dict['weight'] = '0'
            if data_dict.get('price') is None:
                data_dict['price'] = '0'
            data_dict['sum'] = int(data_dict.get('price')) * int(data_dict.get('weight')) + int(data_dict.get(
                'in_china'))

            data_dict['opl'] = sheet[i][11].value
            data.append(data_dict)

    for dictionary in data:
        for key, value in dictionary.items():
            if value is None:
                dictionary[key] = 'Не указано'

    if len(data) > 20:
        big_data = []
        str_to_append = ''
        for i in range(len(data)):
            str_to_append += (f'{i + 1} товар\n'
                              f'Номер товара: {data[i]["no"]}\n'
                              f'Номер авто: {data[i]["no_auto"]}\n'
                              f'Дата выдачи: {data[i]["date"]}\n'
                              f'Трек: {data[i]["trek"]}\n'
                              f'Вес: {data[i]["weight"]}\n'
                              f'Цена: {data[i]["price"]}\n'
                              f'Сумма: {data[i]["sum"]}\n'
                              f'Статус: {data[i]["opl"]}\n\n')
            if i % 20 == 0 and i != 0:
                big_data.append(str_to_append)
                str_to_append = ''
        return big_data

    else:
        small_data = []
        str_to_append = ''
        for i in range(len(data)):
            str_to_append += (f'{i + 1} товар\n'
                              f'Номер товара: {data[i]["no"]}\n'
                              f'Номер авто: {data[i]["no_auto"]}\n'
                              f'Дата выдачи: {data[i]["date"]}\n'
                              f'Трек: {data[i]["trek"]}\n'
                              f'Вес: {data[i]["weight"]}\n'
                              f'Цена: {data[i]["price"]}\n'
                              f'Сумма: {data[i]["sum"]}\n'
                              f'Статус: {data[i]["opl"]}\n\n')
        small_data.append(str_to_append)
        return small_data

# print(data_by_id('75', 'GX-0075', '21', '212'))


#         {'Модель': f'GX-{sheet[i][2].value}','ТРЕК': sheet[i][3].value,
#         'Коробка': sheet[i][4].value, 'Позиция': sheet[i][5].value,
#         'В китае': sheet[i][6].value, 'Вес': sheet[i][7].value,
#         'Цена': sheet[i][8].value, 'Сумма': sheet[i][9].value,
#         'Дата выдачи': sheet[i][10].value, 'Оплачено': sheet[i][11].value,
#         'Примичание': sheet[i][12].value, 'Расчет': sheet[i][13].value,
#         'Полка': sheet[i][14].value}


def data_to_notifications():
    wb_search = load_workbook('data.xlsx')
    sheet = wb_search.active
    gx_lst = []
    trek_lst = []
    for i in range(1, sheet.max_row + 1):
        value = sheet.cell(row=i, column=11).value
        if str(value)[2:10] == datetime.datetime.now().strftime('%y-%m-%d'):
            gx_lst.append(sheet[i][2].value)

    gx_lst = list(set(gx_lst))

    for i in range(1, sheet.max_row + 1):
        value = str(sheet.cell(row=i, column=3).value)
        for n in gx_lst:
            if value == str(n):
                trek_lst.append((n, sheet[i][3].value))
    return trek_lst


def dataXlsx(id1, id2, id3, id4):
    wb_search = load_workbook('data.xlsx')
    sheet = wb_search.active
    data = []
    for i in range(1, sheet.max_row + 1):
        data_dict = []
        value = str(sheet.cell(row=i, column=3).value)
        if id1 == value or id2 == value or id3 == value or id4 == value:
            data_dict.append(id1)
            data_dict.append(sheet[i][0].value)
            try:
                data_dict.append(sheet[i][10].value.strftime('%d.%m.%Y'))
            except AttributeError:
                data_dict.append(None)

            data_dict.append(sheet[i][3].value)

            data_dict.append(sheet[i][7].value)
            data_dict.append(sheet[i][8].value)
            data_dict.append(sheet[i][6].value)
            if data_dict[6] is None:
                data_dict[6] = '0'
            if data_dict[4] is None:
                data_dict[4] = '0'
            if data_dict[5] is None:
                data_dict[5] = '0'
            data_dict.append(int(data_dict[5]) * int(data_dict[4]) + int(data_dict[6]))

            data_dict[-1] = sheet[i][11].value
            data.append(data_dict)
    print(data)
    wb = load_workbook('your.xlsx')
    sheet = wb.active
    a = 0
    for i in range(2, len(data)+2):    
        if a == len(data)+1:
            d = data[a]
            sheet[i][0].value = d[0]
            # print(sheet[i][3].value)
            sheet[i][1].value = d[2]
            sheet[i][2].value = d[3]
            sheet[i][3].value = d[4]
            sheet[i][4].value = d[5]
            sheet[i][5].value = d[6]
            sheet[i][6].value = d[7]
            break
        else:
            d = data[a]
            sheet[i][0].value = d[0]
            sheet[i][1].value = d[2]
            sheet[i][2].value = d[3]    
            sheet[i][3].value = d[4]
            sheet[i][4].value = d[5]
            try:
                if float(d[4]) != 0.0 or float(d[5]) != 0.0:
                    # print(d[4], d[5])
                    sum = (d[4] * d[5])
                    sheet[i][5].value = round(sum, 2)
                else:
                    sheet[i][5].value = 0
            except Exception as ex:
                pass
                
            sheet[i][6].value = d[7]
            a+=1
    wb.save('your.xlsx')
    
            

def clear():
    wb = load_workbook('your.xlsx')
    sheet = wb.active
    for i in range(2, 150):
        for x in range(7):
            sheet[i][x].value = None
    wb.save('your.xlsx')
clear()


